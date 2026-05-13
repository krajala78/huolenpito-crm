from flask import Flask, render_template, request, jsonify, session, send_file
import os, io
import pandas as pd
from datetime import datetime
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── DB driver selection ──────────────────────────────────────────────────────
DATABASE_URL = os.environ.get('DATABASE_URL', '')
if DATABASE_URL:
    import psycopg2, psycopg2.extras
    USE_PG = True
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
else:
    import sqlite3
    USE_PG = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'valueLKV-secret-key-2024-change-in-production')
DB_PATH = os.environ.get('DB_PATH', os.path.join(os.path.dirname(__file__), 'data', 'crm.db'))

KUUKAUDET = ['Tammikuu','Helmikuu','Maaliskuu','Huhtikuu','Toukokuu','Kesäkuu',
             'Heinäkuu','Elokuu','Syyskuu','Lokakuu','Marraskuu','Joulukuu']

EXCEL_TO_DB = {
    'Kohde/Osoite': 'kohde_osoite', 'Omistaja': 'omistaja',
    'Vuokranantajan \nkontakti': 'vuokranantajan_kontakti',
    'Vuokranantajan \nsähköposti': 'vuokranantajan_sahkoposti',
    'Vuokranantajan \npuhelinnumero': 'vuokranantajan_puhelin',
    'Tyyppi': 'tyyppi', 'Koko': 'koko', 'Kaupunki': 'kaupunki',
    'Postinumero': 'postinumero', 'Huolenpito-\nsopimus': 'huolenpitosopimus',
    'Huolen-\npidossa': 'huolenpidossa', 'Vuokrauksessa': 'vuokrauksessa',
    'Vuokravälittäjä': 'vuokravalittaja', 'Vastuuhenkilö': 'vastuuhenkilo',
    'Laskutusperuste sis. alv (€/kk)': 'laskutusperuste',
    'Huolenpidon\nlaskutuksen status': 'laskutuksen_status',
    'Vuokratilitykset': 'vuokratilitykset',
    'Vuokrasopimus alkaen': 'vuokrasopimus_alkaen',
    'Vuokrasopimus päättyy': 'vuokrasopimus_paattyy',
    'Vuokrattu': 'vuokrattu', 'Vuokra-\nmarkkinalla': 'vuokramarkkinalla',
    'Asunnon tila': 'asunnon_tila', 'Vuokralaisen nimi': 'vuokralaisen_nimi',
    'Vuokralaisen puhelinnumero': 'vuokralaisen_puhelin',
    'Vuokralaisen sähköposti': 'vuokralaisen_sahkoposti',
    'Vuokran määrä sop. alkamisessa': 'vuokra_alussa',
    'Vuokran määrä (tänään)': 'vuokra_tanaan',
    'Vesimaksut': 'vesimaksut', 'Muut maksut': 'muut_maksut',
    'Saunamaksut': 'saunamaksut', 'Kokonaisumma': 'kokonaisumma',
    'Vuokravakuus': 'vuokravakuus', 'Vakuuden maksupv.': 'vakuuden_maksupv',
    'Kenen tilillä vakuus': 'kenen_tililla_vakuus',
    'Avaimet luovutettu': 'avaimet_luovutettu',
    'Vesimittari\nluettu': 'vesimittari_luettu',
    'Välitys\nlaskutettu': 'valitys_laskutettu',
    'Välityshinta\nsis. alv': 'valityshinta',
    'Lisätietoja': 'lisatietoja',
}

# ── DB helpers ───────────────────────────────────────────────────────────────
def get_db():
    if USE_PG:
        return psycopg2.connect(DATABASE_URL)
    else:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn

def placeholder():
    return '%s' if USE_PG else '?'

def execute_query(conn, sql, params=None):
    if USE_PG:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params or [])
            return [dict(r) for r in cur.fetchall()]
    else:
        rows = conn.execute(sql, params or []).fetchall()
        return [dict(r) for r in rows]

def execute_one(conn, sql, params=None):
    if USE_PG:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params or [])
            row = cur.fetchone()
            return dict(row) if row else None
    else:
        row = conn.execute(sql, params or []).fetchone()
        return dict(row) if row else None

def execute_scalar(conn, sql, params=None):
    if USE_PG:
        with conn.cursor() as cur:
            cur.execute(sql, params or [])
            row = cur.fetchone()
            return row[0] if row else None
    else:
        row = conn.execute(sql, params or []).fetchone()
        return row[0] if row else None

def execute_write(conn, sql, params=None):
    if USE_PG:
        with conn.cursor() as cur:
            cur.execute(sql, params or [])
            if cur.description:
                row = cur.fetchone()
                return row[0] if row else None
            return None
    else:
        cur = conn.execute(sql, params or [])
        return cur.lastrowid

# ── DB init & migration ──────────────────────────────────────────────────────
def migrate_db():
    """Add new columns to existing tables if they don't exist."""
    new_cols = [
        ('valitys_laskutettu_pvm', 'TEXT'),
        ('takuupalvelu', 'TEXT'),
        ('avainten_lkm', 'INTEGER'),
        ('avainten_luovutettu_lkm', 'INTEGER'),
        ('arkistoitu', 'INTEGER DEFAULT 0'),
    ]
    conn = get_db()
    for col, col_type in new_cols:
        try:
            if USE_PG:
                with conn.cursor() as cur:
                    cur.execute(f'ALTER TABLE properties ADD COLUMN IF NOT EXISTS {col} {col_type}')
                conn.commit()
            else:
                try:
                    conn.execute(f'ALTER TABLE properties ADD COLUMN {col} {col_type}')
                    conn.commit()
                except Exception:
                    pass
        except Exception:
            if USE_PG:
                conn.rollback()
    conn.close()

def init_db():
    if not USE_PG:
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = get_db()
    base_cols = """
        id {pk},
        kohde_osoite TEXT, omistaja TEXT,
        vuokranantajan_kontakti TEXT, vuokranantajan_sahkoposti TEXT,
        vuokranantajan_puhelin TEXT, tyyppi TEXT, koko REAL,
        kaupunki TEXT, postinumero TEXT, huolenpitosopimus TEXT,
        huolenpidossa TEXT, vuokrauksessa TEXT, vuokravalittaja TEXT,
        vastuuhenkilo TEXT, laskutusperuste TEXT, laskutuksen_status TEXT,
        vuokratilitykset TEXT, vuokrasopimus_alkaen TEXT,
        vuokrasopimus_paattyy TEXT, vuokrattu TEXT, vuokramarkkinalla TEXT,
        asunnon_tila TEXT, vuokralaisen_nimi TEXT, vuokralaisen_puhelin TEXT,
        vuokralaisen_sahkoposti TEXT, vuokra_alussa REAL, vuokra_tanaan REAL,
        vesimaksut REAL, muut_maksut TEXT, saunamaksut TEXT,
        kokonaisumma REAL, vuokravakuus REAL, vakuuden_maksupv TEXT,
        kenen_tililla_vakuus TEXT, avaimet_luovutettu TEXT,
        vesimittari_luettu TEXT, valitys_laskutettu TEXT, valityshinta REAL,
        lisatietoja TEXT,
        valitys_laskutettu_pvm TEXT, takuupalvelu TEXT,
        avainten_lkm INTEGER, avainten_luovutettu_lkm INTEGER,
        arkistoitu INTEGER DEFAULT 0,
        luotu TEXT, paivitetty TEXT
    """
    users_cols = """
        id {pk},
        username TEXT UNIQUE NOT NULL, fullname TEXT NOT NULL,
        password_hash TEXT NOT NULL, role TEXT NOT NULL DEFAULT 'user',
        active INTEGER NOT NULL DEFAULT 1, created TEXT NOT NULL
    """
    if USE_PG:
        with conn.cursor() as cur:
            cur.execute(f'CREATE TABLE IF NOT EXISTS properties ({base_cols.format(pk="SERIAL PRIMARY KEY")})')
            cur.execute(f'CREATE TABLE IF NOT EXISTS users ({users_cols.format(pk="SERIAL PRIMARY KEY")})')
        conn.commit()
        count = execute_scalar(conn, 'SELECT COUNT(*) FROM users')
        if count == 0:
            execute_write(conn,
                'INSERT INTO users (username,fullname,password_hash,role,active,created) VALUES (%s,%s,%s,%s,%s,%s)',
                ('admin','Pääkäyttäjä',generate_password_hash('admin123'),'admin',1,datetime.now().strftime('%Y-%m-%d')))
            conn.commit()
    else:
        conn.execute(f'CREATE TABLE IF NOT EXISTS properties ({base_cols.format(pk="INTEGER PRIMARY KEY AUTOINCREMENT")})')
        conn.execute(f'CREATE TABLE IF NOT EXISTS users ({users_cols.format(pk="INTEGER PRIMARY KEY AUTOINCREMENT")})')
        conn.commit()
        count = execute_scalar(conn, 'SELECT COUNT(*) FROM users')
        if count == 0:
            conn.execute('INSERT INTO users (username,fullname,password_hash,role,active,created) VALUES (?,?,?,?,?,?)',
                ('admin','Pääkäyttäjä',generate_password_hash('admin123'),'admin',1,datetime.now().strftime('%Y-%m-%d')))
            conn.commit()
    conn.close()
    migrate_db()

# ── Auth decorators ──────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Kirjautuminen vaaditaan', 'code': 'UNAUTHORIZED'}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Kirjautuminen vaaditaan', 'code': 'UNAUTHORIZED'}), 401
        if session.get('role') != 'admin':
            return jsonify({'error': 'Ei oikeuksia', 'code': 'FORBIDDEN'}), 403
        return f(*args, **kwargs)
    return decorated

# ── Pages ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

# ── Auth API ─────────────────────────────────────────────────────────────────
@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    p = placeholder()
    conn = get_db()
    row = execute_one(conn, f'SELECT * FROM users WHERE username = {p} AND active = 1', (username,))
    conn.close()
    if not row or not check_password_hash(row['password_hash'], password):
        return jsonify({'error': 'Väärä käyttäjätunnus tai salasana'}), 401
    session.permanent = True
    session['user_id'] = row['id']
    session['username'] = row['username']
    session['fullname'] = row['fullname']
    session['role'] = row['role']
    return jsonify({'id': row['id'], 'username': row['username'],
                    'fullname': row['fullname'], 'role': row['role']})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'message': 'Kirjauduttu ulos'})

@app.route('/api/me')
def api_me():
    if 'user_id' not in session:
        return jsonify({'error': 'Ei kirjautunut', 'code': 'UNAUTHORIZED'}), 401
    return jsonify({'id': session['user_id'], 'username': session['username'],
                    'fullname': session['fullname'], 'role': session['role']})

@app.route('/api/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.get_json()
    old_pw = data.get('old_password') or ''
    new_pw = data.get('new_password') or ''
    if len(new_pw) < 4:
        return jsonify({'error': 'Salasanan on oltava vähintään 4 merkkiä'}), 400
    p = placeholder()
    conn = get_db()
    row = execute_one(conn, f'SELECT * FROM users WHERE id = {p}', (session['user_id'],))
    if not row or not check_password_hash(row['password_hash'], old_pw):
        conn.close()
        return jsonify({'error': 'Nykyinen salasana on väärä'}), 400
    execute_write(conn, f'UPDATE users SET password_hash = {p} WHERE id = {p}',
                  (generate_password_hash(new_pw), session['user_id']))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Salasana vaihdettu'})

# ── User management API ──────────────────────────────────────────────────────
@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    conn = get_db()
    rows = execute_query(conn, 'SELECT id,username,fullname,role,active,created FROM users ORDER BY id ASC')
    conn.close()
    return jsonify(rows)

@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    data = request.get_json()
    username = (data.get('username') or '').strip()
    fullname = (data.get('fullname') or '').strip()
    password = data.get('password') or ''
    role = data.get('role', 'user')
    active = 1 if data.get('active', True) else 0
    if not username or not fullname:
        return jsonify({'error': 'Käyttäjätunnus ja nimi ovat pakollisia'}), 400
    if len(password) < 4:
        return jsonify({'error': 'Salasanan on oltava vähintään 4 merkkiä'}), 400
    if role not in ('admin', 'user'):
        return jsonify({'error': 'Virheellinen rooli'}), 400
    p = placeholder()
    conn = get_db()
    if execute_one(conn, f'SELECT id FROM users WHERE username = {p}', (username,)):
        conn.close()
        return jsonify({'error': 'Käyttäjätunnus on jo käytössä'}), 409
    if USE_PG:
        new_id = execute_write(conn,
            'INSERT INTO users (username,fullname,password_hash,role,active,created) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id',
            (username, fullname, generate_password_hash(password), role, active, datetime.now().strftime('%Y-%m-%d')))
    else:
        new_id = execute_write(conn,
            'INSERT INTO users (username,fullname,password_hash,role,active,created) VALUES (?,?,?,?,?,?)',
            (username, fullname, generate_password_hash(password), role, active, datetime.now().strftime('%Y-%m-%d')))
    conn.commit(); conn.close()
    return jsonify({'id': new_id, 'message': 'Käyttäjä luotu'}), 201

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def update_user(user_id):
    data = request.get_json()
    username = (data.get('username') or '').strip()
    fullname = (data.get('fullname') or '').strip()
    role = data.get('role', 'user')
    active = 1 if data.get('active', True) else 0
    new_password = data.get('password') or ''
    if not username or not fullname:
        return jsonify({'error': 'Käyttäjätunnus ja nimi ovat pakollisia'}), 400
    if role not in ('admin', 'user'):
        return jsonify({'error': 'Virheellinen rooli'}), 400
    p = placeholder()
    conn = get_db()
    if execute_one(conn, f'SELECT id FROM users WHERE username = {p} AND id != {p}', (username, user_id)):
        conn.close()
        return jsonify({'error': 'Käyttäjätunnus on jo käytössä'}), 409
    if new_password:
        if len(new_password) < 4:
            conn.close()
            return jsonify({'error': 'Salasanan on oltava vähintään 4 merkkiä'}), 400
        execute_write(conn,
            f'UPDATE users SET username={p},fullname={p},role={p},active={p},password_hash={p} WHERE id={p}',
            (username, fullname, role, active, generate_password_hash(new_password), user_id))
    else:
        execute_write(conn, f'UPDATE users SET username={p},fullname={p},role={p},active={p} WHERE id={p}',
                      (username, fullname, role, active, user_id))
    conn.commit(); conn.close()
    if user_id == session.get('user_id'):
        session['username'] = username; session['fullname'] = fullname; session['role'] = role
    return jsonify({'message': 'Käyttäjä päivitetty'})

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    if user_id == session.get('user_id'):
        return jsonify({'error': 'Et voi poistaa omaa tiliäsi'}), 400
    p = placeholder()
    conn = get_db()
    execute_write(conn, f'DELETE FROM users WHERE id = {p}', (user_id,))
    conn.commit(); conn.close()
    return jsonify({'message': 'Käyttäjä poistettu'})

# ── Stats ────────────────────────────────────────────────────────────────────
@app.route('/api/stats')
@login_required
def stats():
    conn = get_db()
    total = execute_scalar(conn, 'SELECT COUNT(*) FROM properties WHERE arkistoitu = 0')
    vuokrattu = execute_scalar(conn,
        "SELECT COUNT(*) FROM properties WHERE LOWER(vuokrattu) = 'kyllä' AND arkistoitu = 0")
    vapaat = (total or 0) - (vuokrattu or 0)
    vuokramarkkinalla = execute_scalar(conn,
        "SELECT COUNT(*) FROM properties WHERE LOWER(vuokramarkkinalla) = 'kyllä' AND arkistoitu = 0")
    vuokra_sum = execute_scalar(conn,
        "SELECT SUM(kokonaisumma) FROM properties WHERE kokonaisumma IS NOT NULL "
        "AND LOWER(vuokrattu) = 'kyllä' AND arkistoitu = 0") or 0
    huolenpidossa = execute_scalar(conn,
        "SELECT COUNT(*) FROM properties WHERE LOWER(huolenpidossa) = 'kyllä' AND arkistoitu = 0")
    per_vastuuhenkilo = execute_query(conn,
        "SELECT vastuuhenkilo, COUNT(*) as total, "
        "SUM(CASE WHEN LOWER(vuokrattu) = 'kyllä' THEN 1 ELSE 0 END) as vuokrattu "
        "FROM properties WHERE vastuuhenkilo IS NOT NULL AND vastuuhenkilo != '' AND arkistoitu = 0 "
        "GROUP BY vastuuhenkilo ORDER BY total DESC")
    per_tila = execute_query(conn,
        "SELECT COALESCE(asunnon_tila, 'Ei tietoa') as asunnon_tila, COUNT(*) as count "
        "FROM properties WHERE arkistoitu = 0 GROUP BY asunnon_tila ORDER BY count DESC")
    if DATABASE_URL:
        laskutus_sql = ("SELECT SUM(CASE WHEN laskutusperuste ~ '^-?[0-9]+\.?[0-9]*$' "
                        "THEN laskutusperuste::FLOAT8 ELSE 0 END) "
                        "FROM properties WHERE arkistoitu = 0")
    else:
        laskutus_sql = ("SELECT SUM(CAST(laskutusperuste AS REAL)) "
                        "FROM properties WHERE laskutusperuste IS NOT NULL AND arkistoitu = 0")
    laskutus_sum = execute_scalar(conn, laskutus_sql) or 0
    conn.close()
    return jsonify({
        'total': total, 'vuokrattu': vuokrattu, 'vapaat': vapaat,
        'vuokramarkkinalla': vuokramarkkinalla or 0,
        'vuokra_sum': round(float(vuokra_sum), 2),
        'laskutus_sum': round(float(laskutus_sum), 2),
        'huolenpidossa': huolenpidossa or 0,
        'per_vastuuhenkilo': per_vastuuhenkilo,
        'per_tila': per_tila,
    })

# ── Properties ───────────────────────────────────────────────────────────────
@app.route('/api/properties', methods=['GET'])
@login_required
def get_properties():
    search = request.args.get('search', '').strip()
    vuokrattu_f = request.args.get('vuokrattu', '').strip()
    kaupunki_f = request.args.get('kaupunki', '').strip()
    vastuuhenkilo_f = request.args.get('vastuuhenkilo', '').strip()
    arkisto = request.args.get('arkisto', '0').strip()

    p = placeholder()
    like = 'ILIKE' if USE_PG else 'LIKE'
    ark_filter = 'arkistoitu = 1' if arkisto == '1' else 'arkistoitu = 0'
    query = f'SELECT * FROM properties WHERE {ark_filter}'
    params = []

    if search:
        query += (f' AND (kohde_osoite {like} {p} OR omistaja {like} {p}'
                  f' OR vuokralaisen_nimi {like} {p} OR vuokranantajan_kontakti {like} {p}'
                  f' OR kaupunki {like} {p} OR tyyppi {like} {p}'
                  f' OR asunnon_tila {like} {p} OR vastuuhenkilo {like} {p}'
                  f' OR postinumero {like} {p} OR lisatietoja {like} {p})')
        params.extend([f'%{search}%'] * 10)
    if vuokrattu_f:
        query += f' AND LOWER(vuokrattu) = LOWER({p})'; params.append(vuokrattu_f)
    if kaupunki_f:
        query += f' AND kaupunki = {p}'; params.append(kaupunki_f)
    if vastuuhenkilo_f:
        query += f' AND vastuuhenkilo = {p}'; params.append(vastuuhenkilo_f)

    query += ' ORDER BY id ASC'
    conn = get_db()
    rows = execute_query(conn, query, params)
    conn.close()
    return jsonify(rows)

@app.route('/api/properties/<int:prop_id>', methods=['GET'])
@login_required
def get_property(prop_id):
    p = placeholder()
    conn = get_db()
    row = execute_one(conn, f'SELECT * FROM properties WHERE id = {p}', (prop_id,))
    conn.close()
    if row: return jsonify(row)
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/properties', methods=['POST'])
@login_required
def create_property():
    data = request.get_json()
    now = datetime.now().isoformat()
    data['luotu'] = now; data['paivitetty'] = now; data['arkistoitu'] = 0
    data.pop('id', None)
    cols = list(data.keys()); values = [data[k] for k in cols]
    col_str = ', '.join(cols)
    if USE_PG:
        phs = ', '.join(['%s'] * len(cols))
        sql = f'INSERT INTO properties ({col_str}) VALUES ({phs}) RETURNING id'
    else:
        phs = ', '.join(['?'] * len(cols))
        sql = f'INSERT INTO properties ({col_str}) VALUES ({phs})'
    conn = get_db()
    new_id = execute_write(conn, sql, values)
    conn.commit(); conn.close()
    return jsonify({'id': new_id, 'message': 'Luotu onnistuneesti'}), 201

@app.route('/api/properties/<int:prop_id>', methods=['PUT'])
@login_required
def update_property(prop_id):
    data = request.get_json()
    data['paivitetty'] = datetime.now().isoformat()
    data.pop('id', None); data.pop('luotu', None); data.pop('arkistoitu', None)
    if USE_PG:
        set_clause = ', '.join([f'{k} = %s' for k in data.keys()])
        sql = f'UPDATE properties SET {set_clause} WHERE id = %s'
    else:
        set_clause = ', '.join([f'{k} = ?' for k in data.keys()])
        sql = f'UPDATE properties SET {set_clause} WHERE id = ?'
    values = list(data.values()) + [prop_id]
    conn = get_db()
    execute_write(conn, sql, values)
    conn.commit(); conn.close()
    return jsonify({'message': 'Päivitetty onnistuneesti'})

@app.route('/api/properties/<int:prop_id>/archive', methods=['PUT'])
@login_required
def archive_property(prop_id):
    p = placeholder()
    try:
        conn = get_db()
        execute_write(conn, f'UPDATE properties SET arkistoitu = 1, paivitetty = {p} WHERE id = {p}',
                      (datetime.now().isoformat(), prop_id))
        conn.commit(); conn.close()
        return jsonify({'message': 'Kohde arkistoitu'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/properties/<int:prop_id>/restore', methods=['PUT'])
@login_required
def restore_property(prop_id):
    p = placeholder()
    try:
        conn = get_db()
        execute_write(conn, f'UPDATE properties SET arkistoitu = 0, paivitetty = {p} WHERE id = {p}',
                      (datetime.now().isoformat(), prop_id))
        conn.commit(); conn.close()
        return jsonify({'message': 'Kohde palautettu'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/properties/<int:prop_id>', methods=['DELETE'])
@login_required
def delete_property(prop_id):
    p = placeholder()
    try:
        conn = get_db()
        execute_write(conn, f'DELETE FROM properties WHERE id = {p}', (prop_id,))
        conn.commit(); conn.close()
        return jsonify({'message': 'Poistettu pysyvästi'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

NUMERIC_DB_COLS = {
    'koko', 'vuokra_alussa', 'vuokra_tanaan', 'vesimaksut', 'kokonaisumma',
    'vuokravakuus', 'valityshinta', 'avainten_lkm', 'avainten_luovutettu_lkm',
}

def safe_num(val):
    """Convert a value to float/int safely. Returns None if not parseable."""
    if val is None: return None
    if isinstance(val, bool): return None
    if isinstance(val, (int, float)):
        import math
        return None if math.isnan(val) else val
    s = str(val).strip()
    if not s or s.startswith('='): return None   # skip Excel formulas
    # Replace Finnish comma decimal, strip currency symbols and extra text
    s = s.replace(',', '.').replace('€', '').replace(' ', '')
    # Take only the leading numeric part
    import re
    m = re.match(r'^-?[\d]+(?:\.[\d]+)?', s)
    return float(m.group()) if m else None

@app.route('/api/import', methods=['POST'])
@admin_required
def import_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'Tiedostoa ei löydy'}), 400
    file = request.files['file']
    try:
        df = pd.read_excel(io.BytesIO(file.read()))
    except Exception as e:
        return jsonify({'error': f'Tiedoston luku epäonnistui: {str(e)}'}), 400

    count = 0; errors = []

    for idx, row in df.iterrows():
        conn = get_db()
        try:
            data = {}
            for excel_col, db_col in EXCEL_TO_DB.items():
                if excel_col not in df.columns:
                    continue
                val = row[excel_col]
                # NaN / None
                try:
                    is_na = pd.isna(val)
                except (TypeError, ValueError):
                    is_na = False
                if is_na:
                    data[db_col] = None
                elif hasattr(val, 'strftime'):
                    data[db_col] = val.strftime('%Y-%m-%d')
                elif db_col in NUMERIC_DB_COLS:
                    data[db_col] = safe_num(val)
                elif isinstance(val, (int, float)):
                    data[db_col] = val
                else:
                    s = str(val).strip()
                    data[db_col] = s if s else None

            data['luotu']      = datetime.now().isoformat()
            data['paivitetty'] = datetime.now().isoformat()
            data['arkistoitu'] = 0

            cols    = list(data.keys())
            values  = [data[k] for k in cols]
            col_str = ', '.join(cols)
            phs     = ', '.join(['%s' if USE_PG else '?'] * len(cols))
            execute_write(conn, f'INSERT INTO properties ({col_str}) VALUES ({phs})', values)
            conn.commit()
            count += 1
        except Exception as e:
            try: conn.rollback()
            except: pass
            errors.append(f'Rivi {idx + 2}: {str(e).split(chr(10))[0]}')
        finally:
            conn.close()

    result = {'message': f'Tuotu {count} kohdetta', 'count': count}
    if errors: result['errors'] = errors
    return jsonify(result)

@app.route('/api/filters')
@login_required
def get_filters():
    conn = get_db()
    kaupungit = [r['kaupunki'] for r in execute_query(conn,
        'SELECT DISTINCT kaupunki FROM properties WHERE kaupunki IS NOT NULL AND kaupunki != \'\' AND arkistoitu = 0 ORDER BY kaupunki')]
    vastuuhenkilot = [r['vastuuhenkilo'] for r in execute_query(conn,
        'SELECT DISTINCT vastuuhenkilo FROM properties WHERE vastuuhenkilo IS NOT NULL AND vastuuhenkilo != \'\' AND arkistoitu = 0 ORDER BY vastuuhenkilo')]
    conn.close()
    return jsonify({'kaupungit': kaupungit, 'vastuuhenkilot': vastuuhenkilot})

# ── Excel Export ─────────────────────────────────────────────────────────────
@app.route('/api/export')
@login_required
def export_excel():
    ids_param = request.args.get('ids', '').strip()
    conn = get_db()
    if ids_param:
        try:
            ids = [int(i) for i in ids_param.split(',') if i.strip().isdigit()]
        except Exception:
            ids = []
        if ids:
            p = placeholder()
            if USE_PG:
                placeholders = ','.join(['%s'] * len(ids))
            else:
                placeholders = ','.join(['?'] * len(ids))
            rows = execute_query(conn,
                f'SELECT * FROM properties WHERE arkistoitu = 0 AND id IN ({placeholders}) ORDER BY omistaja, kohde_osoite',
                ids)
        else:
            rows = execute_query(conn, 'SELECT * FROM properties WHERE arkistoitu = 0 ORDER BY omistaja, kohde_osoite')
    else:
        rows = execute_query(conn, 'SELECT * FROM properties WHERE arkistoitu = 0 ORDER BY omistaja, kohde_osoite')
    conn.close()
    # Map column names to Finnish
    col_map = {v: k for k, v in EXCEL_TO_DB.items()}
    df = pd.DataFrame(rows)
    df = df.drop(columns=[c for c in ['id','luotu','paivitetty','arkistoitu'] if c in df.columns], errors='ignore')
    df.columns = [col_map.get(c, c) for c in df.columns]
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Huolenpito CRM')
    output.seek(0)
    filename = f'huolenpito_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename)

# ── Tilitysraportti ──────────────────────────────────────────────────────────
@app.route('/api/tilitysraportti')
@login_required
def tilitysraportti():
    vuosi = int(request.args.get('vuosi', datetime.now().year))
    kuukausi = int(request.args.get('kuukausi', datetime.now().month))
    kuukausi_nimi = KUUKAUDET[kuukausi - 1]

    conn = get_db()
    rows = execute_query(conn,
        "SELECT * FROM properties WHERE LOWER(vuokrattu) = 'kyllä' AND arkistoitu = 0 "
        "ORDER BY omistaja, kohde_osoite")
    conn.close()

    wb = Workbook()
    wb.remove(wb.active)

    # ── Kuukausisheet ──
    ws = wb.create_sheet(f'{kuukausi_nimi} {vuosi}')
    headers = ['Osoite','Omistaja','Vuokralainen','Vuokra','Vesimaksu',
               'Muut maksut','Huolenpitomaksut','Muut huomiot','Yhteensä/Tilitettävä summa']
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(bold=True, color='FFFFFF')
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    totals = {4: 0, 5: 0, 6: 0, 7: 0, 9: 0}
    row_num = 2
    alt_fill = PatternFill('solid', fgColor='EBF3FB')

    for prop in rows:
        vuokra = float(prop.get('vuokra_tanaan') or 0)
        vesi = float(prop.get('vesimaksut') or 0)
        muut_raw = str(prop.get('muut_maksut') or '').strip()
        try:
            muut = float(muut_raw) if muut_raw and muut_raw.replace('.','').replace('-','').isdigit() else 0
        except ValueError:
            muut = 0
        lasku_raw = str(prop.get('laskutusperuste') or '').strip()
        try:
            huolenpito = -abs(float(lasku_raw)) if lasku_raw and lasku_raw.replace('.','').replace('-','').isdigit() else 0
        except ValueError:
            huolenpito = 0
        yhteensa = vuokra + vesi + muut + huolenpito
        fill = alt_fill if row_num % 2 == 0 else None
        vals = [prop.get('kohde_osoite',''), prop.get('omistaja',''),
                prop.get('vuokralaisen_nimi',''),
                vuokra or None, vesi or None, muut or None, huolenpito or None, None, yhteensa]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=i, value=v)
            if fill: c.fill = fill
            if i in (4,5,6,7,9) and v is not None:
                c.number_format = '#,##0.00'
        totals[4] += vuokra; totals[5] += vesi; totals[6] += muut
        totals[7] += huolenpito; totals[9] += yhteensa
        row_num += 1

    # Totals row
    tot_font = Font(bold=True)
    tot_fill = PatternFill('solid', fgColor='D6E4F0')
    for col, val in totals.items():
        c = ws.cell(row=row_num, column=col, value=round(val, 2))
        c.font = tot_font; c.fill = tot_fill; c.number_format = '#,##0.00'

    # Empty row then net total
    row_num += 2
    c = ws.cell(row=row_num, column=9, value=round(totals[9], 2))
    c.font = Font(bold=True, color='CC0000'); c.number_format = '#,##0.00'

    # Column widths
    widths = {1:35, 2:25, 3:30, 4:10, 5:10, 6:12, 7:16, 8:14, 9:22}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'

    # Yhteenveto sheet
    ws2 = wb.create_sheet('Yhteenveto ' + str(vuosi))
    ws2.cell(1,1,'Vuokratilitysyhteenveto ' + str(vuosi)).font = Font(bold=True, size=13)
    ws2.cell(3,1,'Kuukausi').font = Font(bold=True)
    ws2.cell(3,2,'Vuokratilitys').font = Font(bold=True)
    ws2.cell(3,3,'Muut tilitykset').font = Font(bold=True)
    ws2.cell(3,4,'Muut huomiot').font = Font(bold=True)
    ws2.cell(4+kuukausi-1, 1, kuukausi_nimi)
    ws2.cell(4+kuukausi-1, 2, round(totals[4], 2)).number_format = '#,##0.00'
    ws2.cell(4+kuukausi-1, 3, round(totals[5]+totals[6], 2)).number_format = '#,##0.00'
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16

    # Logo / website note
    ws2.cell(12, 5, 'www.valuelkv.fi').font = Font(color='888888', italic=True)

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    filename = 'tilitysraportti_' + str(vuosi) + '_' + str(kuukausi).zfill(2) + '.xlsx'
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename)

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_DEBUG', 'false').lower() == 'true')
